"""
write-report.py
Reads a JSON payload from argv[1], writes a formatted .xlsx report.
Sheet 1: Summary
Sheet 2: Full Log
"""
import sys, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Load payload ──────────────────────────────────────────────
with open(sys.argv[1], encoding='utf-8') as f:
    data = json.load(f)

out_path  = data['outPath']
s         = data['summary']
log_lines = data['logLines']
txn       = data.get('txnData')   # optional — present for paygate deposit tests

# ── Strip ANSI codes and illegal XML characters ───────────────
ANSI_RE    = re.compile(r'\x1B\[[0-9;]*[A-Za-z]|\r')
ILLEGAL_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]')

def clean(text):
    text = ANSI_RE.sub('', str(text))
    text = ILLEGAL_RE.sub('', text)
    return text.strip()

log_lines = [clean(line) for line in log_lines if clean(line)]

# ── Colours ───────────────────────────────────────────────────
C_BG_HEADER = '1E2A3A'
C_BG_PASS   = '0D2E22'
C_BG_FAIL   = '2E0F0F'
C_BG_ROW_A  = '0E1218'
C_BG_ROW_B  = '141920'
C_TEXT_LIGHT= 'C9D4E0'
C_TEXT_DIM  = '5A6A7E'
C_ACCENT    = '3B8EE8'
C_GREEN     = '34D399'
C_RED       = 'F87171'
C_AMBER     = 'FBBF24'

def fill(hex_col):
    return PatternFill('solid', start_color=hex_col, fgColor=hex_col)

def font(hex_col, size=11, bold=False, mono=False):
    return Font(name='Courier New' if mono else 'Arial', size=size, bold=bold, color=hex_col)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

wb = Workbook()

# ═══════════════════════════════════════════════════
#  SHEET 1 — SUMMARY
# ═══════════════════════════════════════════════════
ws = wb.active
ws.title = 'Summary'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 36

is_pass = s['result'] == 'PASS'

# Title
ws.merge_cells('A1:B1')
ws['A1'] = '  QA TEST REPORT'
ws['A1'].font      = Font(name='Arial', size=14, bold=True, color=C_ACCENT)
ws['A1'].fill      = fill(C_BG_HEADER)
ws['A1'].alignment = left()
ws.row_dimensions[1].height = 34

# Result banner
ws.merge_cells('A2:B2')
ws['A2'] = '  PASSED' if is_pass else '  FAILED'
ws['A2'].font      = Font(name='Arial', size=13, bold=True, color=C_GREEN if is_pass else C_RED)
ws['A2'].fill      = fill(C_BG_PASS if is_pass else C_BG_FAIL)
ws['A2'].alignment = left()
ws.row_dimensions[2].height = 28

# Spacer
ws.row_dimensions[3].height = 6
ws['A3'].fill = fill(C_BG_ROW_A)
ws['B3'].fill = fill(C_BG_ROW_A)

# Data rows
rows = [
    ('Test',               s['testName']),
    ('Environment',        s['env']),
    ('Date / Time',        s['date']),
    ('Result',             s['result']),
    ('Duration (sec)',     s['durationSec']),
    ('Deposit Amount',     s['depositAmount']),
    ('Withdrawal Amount',  s['withdrawalAmount']),
    ('Members',            s['members']),
]

for i, (label, value) in enumerate(rows):
    r  = i + 4
    bg = C_BG_ROW_A if i % 2 == 0 else C_BG_ROW_B
    ws.row_dimensions[r].height = 22

    ws[f'A{r}'] = f'  {label}'
    ws[f'A{r}'].font      = font(C_TEXT_DIM, size=10, bold=True)
    ws[f'A{r}'].fill      = fill(bg)
    ws[f'A{r}'].alignment = left()

    val_colour = C_TEXT_LIGHT
    if label == 'Result':
        val_colour = C_GREEN if is_pass else C_RED
    elif label == 'Environment':
        val_colour = {'STAGING': C_ACCENT, 'UAT': '6EE7B7', 'PROD': C_AMBER}.get(value, C_TEXT_LIGHT)

    ws[f'B{r}'] = f'  {value}'
    ws[f'B{r}'].font      = font(val_colour, size=11, bold=(label == 'Result'))
    ws[f'B{r}'].fill      = fill(bg)
    ws[f'B{r}'].alignment = left()

# Footer
fr = len(rows) + 4
ws.merge_cells(f'A{fr}:B{fr}')
ws[f'A{fr}'] = '  See "Log Output" sheet for full test log'
ws[f'A{fr}'].font      = font(C_TEXT_DIM, size=9)
ws[f'A{fr}'].fill      = fill(C_BG_HEADER)
ws[f'A{fr}'].alignment = left()
ws.row_dimensions[fr].height = 18

# ── Transaction Details section (paygate deposit tests) ──────────
if txn:
    tr_start = fr + 2  # one blank row gap

    # Section header
    ws.merge_cells(f'A{tr_start}:B{tr_start}')
    ws[f'A{tr_start}'] = '  TRANSACTION DETAILS'
    ws[f'A{tr_start}'].font      = Font(name='Arial', size=12, bold=True, color=C_ACCENT)
    ws[f'A{tr_start}'].fill      = fill(C_BG_HEADER)
    ws[f'A{tr_start}'].alignment = left()
    ws.row_dimensions[tr_start].height = 26

    def _txn_status_colour(status):
        s_lower = str(status).lower()
        if 'approv' in s_lower:  return C_GREEN
        if 'reject' in s_lower:  return C_RED
        if 'timeout' in s_lower: return C_AMBER
        return C_TEXT_LIGHT

    txn_rows = [
        ('Player',             txn.get('player', '—')),
        ('Gateway',            txn.get('gateway', '—')),
        ('Method',             txn.get('method', '—')),
        ('Package',            txn.get('packageName', '—')),
        ('Transaction No',     txn.get('txNo', '—')),
        ('Transaction Date',   txn.get('txDateTime', '—')),
        ('Amount',             txn.get('txAmount', '—')),
        ('Bonus',              txn.get('bonus', '—')),
        ('Transaction Status', txn.get('txStatus', '—')),
        ('Outstanding Total',  txn.get('outstandingTotal', '—')),
        ('Balance Before',     txn.get('balanceBefore', '—')),
        ('Balance After',      txn.get('balanceAfter', '—')),
        ('Rollover Before',    txn.get('rolloverBefore', '—')),
        ('Rollover After',     txn.get('rolloverAfter', '—')),
        ('Target Before',      txn.get('targetBefore', '—')),
        ('Target After',       txn.get('targetAfter', '—')),
    ]

    for i, (label, value) in enumerate(txn_rows):
        r  = tr_start + 1 + i
        bg = C_BG_ROW_A if i % 2 == 0 else C_BG_ROW_B
        ws.row_dimensions[r].height = 22

        ws[f'A{r}'] = f'  {label}'
        ws[f'A{r}'].font      = font(C_TEXT_DIM, size=10, bold=True)
        ws[f'A{r}'].fill      = fill(bg)
        ws[f'A{r}'].alignment = left()

        val_col = C_TEXT_LIGHT
        if label == 'Transaction Status':
            val_col = _txn_status_colour(value)
        elif label in ('Balance After', 'Rollover After', 'Target After'):
            val_col = C_GREEN if value != '—' else C_TEXT_DIM
        elif label in ('Balance Before', 'Rollover Before', 'Target Before'):
            val_col = C_AMBER

        ws[f'B{r}'] = f'  {value}'
        ws[f'B{r}'].font      = font(val_col, size=11, bold=(label == 'Transaction Status'))
        ws[f'B{r}'].fill      = fill(bg)
        ws[f'B{r}'].alignment = left()

ws.sheet_properties.tabColor = C_ACCENT if is_pass else 'F87171'

# ═══════════════════════════════════════════════════
#  SHEET 2 — LOG OUTPUT
# ═══════════════════════════════════════════════════
wl = wb.create_sheet('Log Output')
wl.sheet_view.showGridLines = False
wl.column_dimensions['A'].width = 6
wl.column_dimensions['B'].width = 120

# Header
wl.merge_cells('A1:B1')
wl['A1'] = '  FULL TEST LOG'
wl['A1'].font      = Font(name='Arial', size=12, bold=True, color=C_ACCENT)
wl['A1'].fill      = fill(C_BG_HEADER)
wl['A1'].alignment = left()
wl.row_dimensions[1].height = 28

# Column headers
wl['A2'] = '#'
wl['B2'] = 'Output'
for col in ['A', 'B']:
    wl[f'{col}2'].font      = font(C_TEXT_DIM, size=9, bold=True, mono=True)
    wl[f'{col}2'].fill      = fill('1B2230')
    wl[f'{col}2'].alignment = center() if col == 'A' else left()
wl.row_dimensions[2].height = 18

# Log lines
for i, line in enumerate(log_lines):
    r  = i + 3
    bg = C_BG_ROW_A if i % 2 == 0 else C_BG_ROW_B

    wl[f'A{r}'] = i + 1
    wl[f'A{r}'].font      = font(C_TEXT_DIM, size=9, mono=True)
    wl[f'A{r}'].fill      = fill(bg)
    wl[f'A{r}'].alignment = center()

    txt = line
    if 'PASSED' in txt or 'passed' in txt.lower():
        col = C_GREEN
    elif 'FAILED' in txt or 'failed' in txt.lower() or 'error' in txt.lower():
        col = C_RED
    elif txt.startswith('>>'):
        col = '90CDF4'
    elif '=' in txt or '-' in txt * (len(txt) > 5 and txt.count('-') > 5):
        col = C_TEXT_DIM
    else:
        col = C_TEXT_LIGHT

    wl[f'B{r}'] = txt
    wl[f'B{r}'].font      = font(col, size=10, mono=True)
    wl[f'B{r}'].fill      = fill(bg)
    wl[f'B{r}'].alignment = left()
    wl.row_dimensions[r].height = 16

wl.sheet_properties.tabColor = '2D3748'

# ═══════════════════════════════════════════════════
#  SHEET 3 — SCREENSHOTS
# ═══════════════════════════════════════════════════
screenshot_paths = data.get('screenshotPaths', [])

if screenshot_paths:
    import os
    from openpyxl.drawing.image import Image as XLImage
    try:
        from PIL import Image as PILImage
        HAS_PIL = True
    except Exception:
        HAS_PIL = False

    ws3 = wb.create_sheet('Screenshots')
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 16   # label column
    ws3.column_dimensions['B'].width = 100  # image column

    # Header
    ws3.merge_cells('A1:B1')
    ws3['A1'] = '  TEST SCREENSHOTS'
    ws3['A1'].font      = Font(name='Arial', size=12, bold=True, color=C_ACCENT)
    ws3['A1'].fill      = fill(C_BG_HEADER)
    ws3['A1'].alignment = left()
    ws3.row_dimensions[1].height = 28

    current_row = 2
    IMG_HEIGHT_PX = 360   # display height per screenshot
    ROW_HEIGHT_PT = IMG_HEIGHT_PX * 0.75  # approx pt conversion

    for i, item in enumerate(screenshot_paths):
        label     = item.get('label', f'Step {i+1}')
        img_path  = item.get('path', '')

        if not img_path or not os.path.exists(img_path):
            continue

        # Label row
        bg = C_BG_ROW_A if i % 2 == 0 else C_BG_ROW_B
        ws3.merge_cells(f'A{current_row}:B{current_row}')
        ws3[f'A{current_row}'] = f'  {label}'
        ws3[f'A{current_row}'].font      = font(C_ACCENT, size=11, bold=True)
        ws3[f'A{current_row}'].fill      = fill('1B2230')
        ws3[f'A{current_row}'].alignment = left()
        ws3.row_dimensions[current_row].height = 20
        current_row += 1

        # Image row
        img_row = current_row
        ws3.row_dimensions[img_row].height = ROW_HEIGHT_PT

        try:
            img = XLImage(img_path)
            # Use PIL for reliable dimension reading
            if HAS_PIL:
                with PILImage.open(img_path) as pil_img:
                    orig_w, orig_h = pil_img.size
            else:
                orig_w = img.width or 1920
                orig_h = img.height or 1080
            if orig_w > 900:
                scale = 900 / orig_w
                img.width  = int(orig_w * scale)
                img.height = int(orig_h * scale)
            ws3.add_image(img, f'B{img_row}')
        except Exception as e:
            ws3[f'B{img_row}'] = f'[Image error: {e}]'
            ws3[f'B{img_row}'].font = font(C_RED, size=10)

        current_row += int(ROW_HEIGHT_PT / 15) + 1  # advance rows to clear image

    ws3.sheet_properties.tabColor = '6366f1'

# ── Save (atomic: write to .tmp then rename so a crash never corrupts the final file) ──
import os as _os
tmp_path = out_path + '.tmp'
wb.save(tmp_path)
_os.replace(tmp_path, out_path)