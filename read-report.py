"""
read-report.py
Reads the Summary sheet from a QA report xlsx and prints JSON to stdout.
Used by server.js /reports/preview endpoint.
"""
import sys, json, re
from openpyxl import load_workbook

ANSI_RE    = re.compile(r'\x1B\[[0-9;]*[A-Za-z]|\r')
ILLEGAL_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]')

def clean(val):
    if val is None:
        return ''
    text = ANSI_RE.sub('', str(val))
    text = ILLEGAL_RE.sub('', text)
    return text.strip()

try:
    wb = load_workbook(sys.argv[1], data_only=True)
except Exception as e:
    print(json.dumps({'error': f'Corrupted report file: {e}', 'summary': {}, 'logLines': [], 'totalLines': 0}))
    sys.exit(0)

ws = wb['Summary']

# Rows 4-11 are the data rows (label in col A, value in col B)
# Row 1 = title, Row 2 = result banner, Row 3 = spacer
summary = {}
label_map = {
    'Test':               'testName',
    'Environment':        'env',
    'Date / Time':        'date',
    'Result':             'result',
    'Duration (sec)':     'durationSec',
    'Deposit Amount':     'depositAmount',
    'Withdrawal Amount':  'withdrawalAmount',
    'Members':            'members',
}

for row in ws.iter_rows(min_row=4, max_row=20, min_col=1, max_col=2, values_only=True):
    label = clean(row[0])
    value = clean(row[1])
    if label in label_map:
        summary[label_map[label]] = value

# Also read log lines from Log Output sheet
log_lines = []
if 'Log Output' in wb.sheetnames:
    wl = wb['Log Output']
    for row in wl.iter_rows(min_row=3, max_row=wl.max_row, min_col=2, max_col=2, values_only=True):
        val = clean(row[0])
        if val:
            log_lines.append(val)

print(json.dumps({ 'summary': summary, 'logLines': log_lines, 'totalLines': len(log_lines) }))
